"""
OpenCV와 pyzbar를 사용하는 바코드 스캐너 유틸리티.

요약:
- 카메라에서 프레임을 읽어 pyzbar로 바코드/QR을 디코딩
- 바코드 위치에 사각형/폴리곤을 그려 화면에 표시
- 새로운 코드가 감지되면 콘솔에 출력하고(옵션) 비프음 재생

사용법: `python scanner.py [CAM_INDEX] [beep_flag] [--debug]`

변경 및 추가된 기능(요약):
- `--debug` 플래그 추가: 각 프레임에 대해 간단한 로그 출력 (frame, decoded count)
- 바코드 표시 라벨은 기본적으로 `{TYPE}: {DATA}` 형식으로 화면에 출력
- 매핑 기능을 위해 `mappings.json`을 추가할 수 있음(현재 워크스페이스에 파일이 존재하면 자동 로드).
    - 예: `mappings.json`에 "3009P": "21008고진우"를 넣으면, CODE128의 값 `3009P` 감지 시 화면과 콘솔에 `21008고진우`로 표시됩니다.
- 편의 실행 스크립트 추가:
    - `run_scanner.bat` — 워크스페이스의 고정 Python 경로로 `scanner.py`를 실행
    - `run_scanner.ps1` — PowerShell용 런처
- VS Code 워크스페이스 설정 추가:
    - `.vscode/settings.json` — `python.defaultInterpreterPath`와 (선택적으로) `Code Runner` 실행기 고정
    - `.vscode/launch.json` — Run/Debug 시 워크스페이스 인터프리터 사용

현재 추가된 파일(워크스페이스 기준):c
- `mappings.json` (바코드->표시 문자열 매핑)
- `run_scanner.bat`, `run_scanner.ps1` (런처)
- `.vscode/settings.json`, `.vscode/launch.json` (워크스페이스 실행 환경 고정)

테스트/실행 팁:
- 터미널에서 항상 동일한 파이썬 인터프리터를 사용하세요:
    `python -c "import sys; print(sys.executable)"`
- 패키지가 설치된 파이썬으로 실행:
    `python scanner.py` 또는 `.\run_scanner.bat`

"""

import sys
import time
# Windows용 비프(선택적); 불가능하면 비활성화
try:
    import winsound
except Exception:
    winsound = None

# OpenCV: 카메라 캡처 및 화면 표시
import cv2
import re
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
 

# pyzbar: zbar 라이브러리 래퍼 — 바코드 디코딩
from pyzbar import pyzbar


# 동일 폴더의 mappings.json에서 키→표시 문자열 매핑을 로드합니다.
def load_mappings():
    p = Path(__file__).parent / 'mappings.json'
    if p.exists():
        try:
            return json.loads(p.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}

MAPPINGS = load_mappings()

EXCEL_PATH = Path(__file__).parent / "scans.xlsx"

def init_excel(path=EXCEL_PATH):
    """엑셀 파일이 없으면 만들고, 헤더를 넣습니다."""
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Scans"
        ws.append(["timestamp", "type", "data", "label"])
        wb.save(path)

def append_excel(ts, typ, data, label, path=EXCEL_PATH):
    """엑셀에 한 줄 추가합니다."""
    wb = load_workbook(path)
    ws = wb.active
    ws.append([ts, typ, data, label])
    wb.save(path)

def reset_excel(path=EXCEL_PATH):
    """엑셀을 완전 초기화(헤더만 남김)합니다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scans"
    ws.append(["timestamp", "type", "data", "label"])
    wb.save(path)



def decode_and_annotate(frame):
    """
    프레임에서 바코드/QR을 디코딩하고 시각적 어노테이션을 추가합니다.

    동작:
    - `pyzbar.decode`로 심볼을 검색
    - 발견된 각 심볼에 대해 사각형과 폴리곤을 그림
    - 심볼 데이터와 타입을 문자열로 디코딩해 프레임에 텍스트로 표시
    - (type, data) 튜플 리스트 반환
    """
    barcodes = pyzbar.decode(frame)
    results = []
    for b in barcodes:
        # 바운딩 박스 및 폴리곤 그리기
        x, y, w, h = b.rect
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
        pts = b.polygon
        if pts:
            pts = [(p.x, p.y) for p in pts]
            for i in range(len(pts)):
                cv2.line(frame, pts[i], pts[(i + 1) % len(pts)], (255, 0, 0), 2)

        # 데이터 디코딩 및 화면에 표시
        data = b.data.decode('utf-8', errors='replace')
        typ = b.type

        # CODE128이고 4자리 숫자+영문 1자 패턴이면 매핑을 시도합니다
        label = f"{typ}: {data}"
        if typ == 'CODE128' and re.match(r'^\d{4}[A-Za-z]$', data):
            mapped = MAPPINGS.get(data)
            if mapped:
                label = mapped

        cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        # 이후 출력용으로 라벨을 함께 반환합니다
        results.append((typ, data, label))
    return frame, results


def main(camera_index=0, beep=True, debug=False):
    """
    메인 루프: 카메라에서 프레임을 읽고 디코딩/어노테이션을 수행합니다.

    옵션:
    - `camera_index`: 사용할 카메라 인덱스 (기본 0)
    - `beep`: 새로운 코드 감지 시 윈도우 비프음을 사용할지 여부
    - `debug`: 각 프레임에 대해 간단한 로그 출력

    동작:
    - 카메라가 열리지 않으면 종료
    - 각 프레임을 `decode_and_annotate`로 처리
    - 새로 감지된 데이터는 `seen` 집합으로 중복 방지 후 콘솔에 출력
    - 프레임을 윈도우에 표시하고 `Esc` 또는 `q`로 종료
    """
    cap = cv2.VideoCapture(camera_index)
    if not cap.isOpened():
        print(f"카메라({camera_index})를 열 수 없습니다.")
        return

    seen = set()
    frame_count = 0
    init_excel()
    try:
        while True:
            # 프레임 읽기
            ret, frame = cap.read()
            frame_count += 1
            if not ret:
                break

            # 디코딩 및 어노테이션
            frame, results = decode_and_annotate(frame)
            if debug:
                print(f"frame={frame_count} ret={ret} decoded={len(results)}")

            # 새로 감지된 코드 처리
            # results는 이제 (typ, data, label) 형식입니다
            for typ, data, label in results:
                if label not in seen:
                    seen.add(label)
                    ts = time.strftime('%Y-%m-%d %H:%M:%S')
                    print(f"[{ts}] {label}")
                    mapped = (typ == 'CODE128'and re.match(r'^\d{4}[A-Za-z]$', data)and (data in MAPPINGS))

                    if mapped:
                        append_excel(ts, typ, data, label)  # label은 매핑된 이름

                    if beep and winsound:
                        try:
                            winsound.Beep(1000, 120)
                        except Exception:
                            pass

            # 화면에 표시 및 키 이벤트 처리
            cv2.imshow('Barcode Scanner', frame)
            key = cv2.waitKey(1) & 0xFF
            if key == ord('c'):  # c 키 → 스캔 기록 초기화
                seen.clear()
                reset_excel()
                print("스캔 기록이 초기화되었습니다.")
            elif key == 27 or key == ord('q'):  # ESC 또는 q → 종료
                break

    finally:    
        cap.release()
        cv2.destroyAllWindows()


if __name__ == '__main__':
    cam = 0
    beep_flag = True
    debug_flag = False
    if len(sys.argv) > 1:
        try:
            cam = int(sys.argv[1])
        except Exception:
            cam = 0
    if len(sys.argv) > 2:
        if sys.argv[2].lower() in ('0', 'false', 'no'):
            beep_flag = False
    # '--debug' 또는 '-d' 플래그도 지원합니다(어떤 위치에서도 가능)
    if '--debug' in sys.argv or '-d' in sys.argv:
        debug_flag = True
    main(camera_index=cam, beep=beep_flag, debug=debug_flag)